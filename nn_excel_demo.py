import os
import pathlib
from typing import List, Tuple

import numpy as np


def ensure_dir(p: str):
    pathlib.Path(p).mkdir(parents=True, exist_ok=True)


QD_BASE = "https://storage.googleapis.com/quickdraw_dataset/full/numpy_bitmap"
CATEGORIES = [
    ("cat", f"{QD_BASE}/cat.npy"),
    ("house", f"{QD_BASE}/house.npy"),
    ("ladder", f"{QD_BASE}/ladder.npy"),
    ("sun", f"{QD_BASE}/sun.npy"),
    ("tree", f"{QD_BASE}/tree.npy"),
    ("door", f"{QD_BASE}/door.npy"),

    
]


def download_quickdraw(data_dir: str, max_per_class: int = 30000):
    import requests

    ensure_dir(data_dir)
    for name, url in CATEGORIES:
        out_path = os.path.join(data_dir, f"{name}.npy")
        if os.path.exists(out_path):
            print(f"[data] Found cached {out_path}")
            continue
        print(f"[data] Downloading {name} from {url}")
        with requests.get(url, stream=True, timeout=60) as r:
            r.raise_for_status()
            with open(out_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=1 << 20):
                    if chunk:
                        f.write(chunk)
        print(f"[data] Saved {out_path}")


def load_dataset(data_dir: str, n_per_class: int = 30000) -> Tuple[np.ndarray, np.ndarray, List[str]]:
    X_list = []
    y_list = []
    label_names = []
    for label_idx, (name, _) in enumerate(CATEGORIES):
        path = os.path.join(data_dir, f"{name}.npy")
        arr = np.load(path)
        # Use first n_per_class samples
        arr = arr[:n_per_class].astype(np.float32)
        X_list.append(arr)
        y_list.append(np.full((arr.shape[0],), label_idx, dtype=np.int64))
        label_names.append(name)

    X = np.concatenate(X_list, axis=0)
    y = np.concatenate(y_list, axis=0)
    # Scale to [0,1]
    X /= 255.0
    # Shuffle
    rng = np.random.default_rng(42)
    idx = rng.permutation(X.shape[0])
    X = X[idx]
    y = y[idx]
    return X, y, label_names


class TinyNN:
    def __init__(self, in_dim=784, hidden=32, out_dim=5, seed=0):
        rng = np.random.default_rng(seed)
        # He init for ReLU
        self.W1 = rng.normal(0, np.sqrt(2.0 / in_dim), size=(in_dim, hidden)).astype(np.float32)
        self.b1 = np.zeros((hidden,), dtype=np.float32)
        # Xavier/Glorot for output
        self.W2 = rng.normal(0, np.sqrt(1.0 / hidden), size=(hidden, out_dim)).astype(np.float32)
        self.b2 = np.zeros((out_dim,), dtype=np.float32)
        # Adam state
        self.mW1 = np.zeros_like(self.W1)
        self.vW1 = np.zeros_like(self.W1)
        self.mb1 = np.zeros_like(self.b1)
        self.vb1 = np.zeros_like(self.b1)
        self.mW2 = np.zeros_like(self.W2)
        self.vW2 = np.zeros_like(self.W2)
        self.mb2 = np.zeros_like(self.b2)
        self.vb2 = np.zeros_like(self.b2)
        self.t = 0

    @staticmethod
    def relu(x):
        return np.maximum(x, 0)

    @staticmethod
    def softmax_logits(logits):
        z = logits - logits.max(axis=1, keepdims=True)
        e = np.exp(z)
        return e / e.sum(axis=1, keepdims=True)

    def forward(self, X):
        h_pre = X @ self.W1 + self.b1
        h = self.relu(h_pre)
        logits = h @ self.W2 + self.b2
        return h_pre, h, logits

    def loss_and_grads(self, X, y):
        # Forward
        h_pre, h, logits = self.forward(X)
        probs = self.softmax_logits(logits)
        # Cross-entropy loss
        N = X.shape[0]
        loss = -np.log(probs[np.arange(N), y] + 1e-12).mean()
        # Gradients
        dlogits = probs
        dlogits[np.arange(N), y] -= 1
        dlogits /= N
        dW2 = h.T @ dlogits
        db2 = dlogits.sum(axis=0)
        dh = dlogits @ self.W2.T
        dh_pre = dh * (h_pre > 0)
        dW1 = X.T @ dh_pre
        db1 = dh_pre.sum(axis=0)
        return loss, (dW1, db1, dW2, db2)

    def step_adam(self, grads, lr=1e-3, beta1=0.9, beta2=0.999, eps=1e-8):
        dW1, db1, dW2, db2 = grads
        self.t += 1
        def update(param, grad, m, v):
            m[:] = beta1 * m + (1 - beta1) * grad
            v[:] = beta2 * v + (1 - beta2) * (grad * grad)
            m_hat = m / (1 - beta1 ** self.t)
            v_hat = v / (1 - beta2 ** self.t)
            param[:] = param - lr * m_hat / (np.sqrt(v_hat) + eps)
        update(self.W1, dW1, self.mW1, self.vW1)
        update(self.b1, db1, self.mb1, self.vb1)
        update(self.W2, dW2, self.mW2, self.vW2)
        update(self.b2, db2, self.mb2, self.vb2)

    def predict(self, X):
        _, _, logits = self.forward(X)
        return logits.argmax(axis=1)


def train_model(X, y, max_epochs=30, target_acc=0.95, batch_size=256, lr=1e-3, val_split=0.15, seed=0, min_epochs=10):
    rng = np.random.default_rng(seed)
    N = X.shape[0]
    # Split
    n_val = int(N * val_split)
    X_val = X[:n_val]
    y_val = y[:n_val]
    X_tr = X[n_val:]
    y_tr = y[n_val:]

    model = TinyNN(in_dim=X.shape[1], hidden=32, out_dim=6, seed=seed)

    best = {
        "epoch": -1,
        "acc": 0.0,
        "W1": model.W1.copy(),
        "b1": model.b1.copy(),
        "W2": model.W2.copy(),
        "b2": model.b2.copy(),
    }

    for epoch in range(1, max_epochs + 1):
        # Shuffle train
        idx = rng.permutation(X_tr.shape[0])
        X_tr = X_tr[idx]
        y_tr = y_tr[idx]

        # Mini-batch
        losses = []
        for i in range(0, X_tr.shape[0], batch_size):
            xb = X_tr[i:i + batch_size]
            yb = y_tr[i:i + batch_size]
            loss, grads = model.loss_and_grads(xb, yb)
            losses.append(loss)
            model.step_adam(grads, lr=lr)

        # Eval
        with np.errstate(over='ignore'):
            pred = model.predict(X_val)
        acc = (pred == y_val).mean()
        mean_loss = float(np.mean(losses))
        print(f"[train] epoch {epoch:3d} | loss {mean_loss:.4f} | val acc {acc*100:.2f}%")

        # Track best
        if acc > best["acc"]:
            best.update(epoch=epoch, acc=float(acc), W1=model.W1.copy(), b1=model.b1.copy(), W2=model.W2.copy(), b2=model.b2.copy())

        # Only allow early stopping after minimum epochs and if target accuracy is reached
        if epoch >= min_epochs and acc >= target_acc:
            print(f"[train] Early stopping at epoch {epoch} (val acc {acc*100:.2f}%)")
            break

        # Simple LR schedule bump if plateau
        if epoch % 50 == 0:
            lr *= 0.5

    # Load best
    model.W1[:] = best["W1"]
    model.b1[:] = best["b1"]
    model.W2[:] = best["W2"]
    model.b2[:] = best["b2"]
    print(f"[train] Best val acc {best['acc']*100:.2f}% at epoch {best['epoch']}")
    return model


# Excel export utilities
def excel_col_letter(n: int) -> str:
    # 1-indexed
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def write_weights_sheet(wb, W1, b1, W2, b2, sheet_name="Weights"):

    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = "CCCCCC"

    # Layout:
    # W1: A1:AF784 (784 rows x 32 cols)
    # b1: row 785 across A:AF
    # gap rows 786-789
    # W2: A790:E821 (32 rows x 5 cols)
    # b2: row 822 across A:E
    print("[excel] Writing W1 (784x32) ...")
    for r in range(W1.shape[0]):
        row_vals = W1[r]
        for c in range(W1.shape[1]):
            ws.cell(row=r + 1, column=c + 1, value=float(row_vals[c]))
    print("[excel] Writing b1 (32) ...")
    for c in range(b1.shape[0]):
        ws.cell(row=785, column=c + 1, value=float(b1[c]))

    print("[excel] Writing W2 (32x5) ...")
    w2_row0 = 790
    for r in range(W2.shape[0]):
        for c in range(W2.shape[1]):
            ws.cell(row=w2_row0 + r, column=c + 1, value=float(W2[r, c]))
    print("[excel] Writing b2 (5) ...")
    for c in range(b2.shape[0]):
        ws.cell(row=822, column=c + 1, value=float(b2[c]))

    # Hide the sheet to keep the demo tidy
    ws.sheet_state = "hidden"
    return ws


def add_data_validation_binary(ws, cell_range: str):
    from openpyxl.worksheet.datavalidation import DataValidation
    # Use a dropdown list for faster drawing with the mouse
    # Set showDropDown=False for cleaner interface - users can still type 0/1 directly
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=False, showDropDown=True)
    dv.error = "Only 0 or 1 allowed"
    dv.errorTitle = "Invalid Input"
    dv.prompt = "Click dropdown or type 0 (erase) or 1 (paint)"
    dv.promptTitle = "Drawing Tool"
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def add_grid_borders(ws, cell_range: str):
    from openpyxl.styles import Border, Side
    from openpyxl.utils import range_boundaries
    
    # Add thin borders to make grid visible
    thin_border = Border(
        left=Side(style='thin', color='808080'),
        right=Side(style='thin', color='808080'),
        top=Side(style='thin', color='808080'),
        bottom=Side(style='thin', color='808080')
    )
    
    # Parse the range and apply borders
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border


def add_conditional_formatting_grid(ws, cell_range: str):
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    
    # Black fill when cell > 0 (painted pixels, treat any non-zero as painted)
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    # Much darker gray fill when cell == 0 (background pixels) for better contrast
    dark_gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Use a relative formula starting from the top-left cell
    top_left = cell_range.split(":")[0]
    
    # Rule for painted pixels (value > 0) - black fill
    black_rule = FormulaRule(formula=[f"{top_left}>0"], stopIfTrue=True, fill=black_fill)
    ws.conditional_formatting.add(cell_range, black_rule)
    
    # Rule for background pixels (value = 0) - dark gray fill
    gray_rule = FormulaRule(formula=[f"{top_left}=0"], stopIfTrue=False, fill=dark_gray_fill)
    ws.conditional_formatting.add(cell_range, gray_rule)
    
    # Add grid borders for better visibility
    add_grid_borders(ws, cell_range)


def add_prediction_formatting(ws):
    from openpyxl.formatting.rule import DataBarRule, FormulaRule
    from openpyxl.styles import PatternFill

    # Data bar on AF2:AF6
    ws.conditional_formatting.add("AF2:AF6", DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color='63BE7B', showValue="True"))

    # Highlight top prediction in AF and AE columns
    yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    # openpyxl doesn't apply font via rule directly; use fill for highlight
    ws.conditional_formatting.add("AF2:AF6", FormulaRule(formula=["AF2=MAX($AF$2:$AF$6)"], stopIfTrue=True, fill=yellow))
    ws.conditional_formatting.add("AE2:AE6", FormulaRule(formula=["AF2=MAX($AF$2:$AF$6)"], stopIfTrue=True, fill=yellow))


def build_draw_sheet(wb, sheet_name: str, labels: List[str], sample_grid: np.ndarray = None):

    ws = wb.create_sheet(title=sheet_name)

    # 28x28 grid in A1:AB28
    grid_rows, grid_cols = 28, 28
    grid_top_left_col = 1  # A
    grid_top_left_row = 1
    for r in range(grid_rows):
        for c in range(grid_cols):
            val = 0
            if sample_grid is not None:
                val = int(sample_grid[r, c])
            cell = ws.cell(row=grid_top_left_row + r, column=grid_top_left_col + c, value=val)
            # Hide the numeric value to make the pixel art clearer
            cell.number_format = ';;;'
    # Make cells square and larger for better mouse painting
    for c in range(1, 28 + 1):
        ws.column_dimensions[excel_col_letter(c)].width = 3.0  # Slightly wider for better visibility
    for r in range(1, 28 + 1):
        ws.row_dimensions[r].height = 18  # Taller for better click targets

    # Data validation for binary input
    add_data_validation_binary(ws, "A1:AB28")
    add_conditional_formatting_grid(ws, "A1:AB28")

    # Clear toggle at AD2 (0/1)
    ws["AD1"] = "Clear (0/1)"
    ws["AD2"] = 0
    add_data_validation_binary(ws, "AD2")

    # Flattened vector in AC1:AC784 = (1-$AD$2) * --(INDEX($A$1:$AB$28, r, c)<>0)
    # We'll write an explicit INDEX reference per row for clarity and performance
    flat_col = "AC"
    k = 0
    for rr in range(1, 28 + 1):
        for cc in range(1, 28 + 1):
            k += 1
            cell = f"{flat_col}{k}"
            # Allow weighted input but cap at 3 to prevent extreme activations while preserving emphasis
            ws[cell] = f"=(1-$AD$2)*MIN(3, MAX(0, INDEX($A$1:$AB$28,{rr},{cc})))"

    # Hidden layer outputs in AI1:AI32
    # Formula: =MAX(0, SUMPRODUCT($AC$1:$AC$784, Weights!<col>1:<col>784) + Weights!<col>$785)
    for j in range(32):
        out_cell = ws.cell(row=1 + j, column=35)  # AI column = 35
        col_letter = excel_col_letter(1 + j)  # Weights! columns A..AF
        out_cell.value = (
            f"=MAX(0, SUMPRODUCT($AC$1:$AC$784, Weights!${col_letter}$1:${col_letter}$784) + Weights!${col_letter}$785)"
        )

    # Logits in AJ1:AJ5
    # Formula: =SUMPRODUCT($AI$1:$AI$32, Weights!<col>790:<col>821) + Weights!<col>$822
    for k_cls in range(5):
        out_cell = ws.cell(row=1 + k_cls, column=36)  # AJ
        col_letter = excel_col_letter(1 + k_cls)  # A..E for W2
        out_cell.value = (
            f"=SUMPRODUCT($AI$1:$AI$32, Weights!${col_letter}$790:${col_letter}$821) + Weights!${col_letter}$822"
        )

    # Probabilities in AK1:AK5: =EXP(AJr)/SUMPRODUCT(EXP($AJ$1:$AJ$5))
    for k_cls in range(5):
        out_cell = ws.cell(row=1 + k_cls, column=37)  # AK
        row = 1 + k_cls
        out_cell.value = f"=EXP(AJ{row})/SUMPRODUCT(EXP($AJ$1:$AJ$5))"

    # Prediction panel in AE-AG
    ws["AE1"] = "Category"
    ws["AF1"] = "Probability"
    ws["AG1"] = "Bar"
    for i, name in enumerate(labels):
        ws.cell(row=2 + i, column=31, value=name)  # AE
        # Map to AK1..AK5
        ws.cell(row=2 + i, column=32, value=f"=AK{1 + i}")  # AF
        ws.cell(row=2 + i, column=33, value=f"=AF{2 + i}")  # AG mirrors value to show data bar

    # Conditional formatting for prediction panel
    add_prediction_formatting(ws)


    # Keep view tidy
    ws.freeze_panes = "AC1"

    return ws





def add_diverging_scale(ws, cell_range: str):
    from openpyxl.formatting.rule import ColorScaleRule
    # Blue (neg) -> White (0) -> Red (pos)
    rule = ColorScaleRule(
        start_type='num', start_value=-1, start_color='2C7BB6',
        mid_type='num', mid_value=0, mid_color='FFFFFF',
        end_type='num', end_value=1, end_color='D7191C'
    )
    ws.conditional_formatting.add(cell_range, rule)


def add_red_heat(ws, cell_range: str):
    from openpyxl.formatting.rule import ColorScaleRule
    # White (low) -> Red (high)
    rule = ColorScaleRule(
        start_type='min', start_color='FFFFFF',
        end_type='max', end_color='FF5252'
    )
    ws.conditional_formatting.add(cell_range, rule)


def add_int_range_validation(ws, cell: str, low: int, high: int):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="whole", operator="between", formula1=str(low), formula2=str(high), allow_blank=False)
    dv.error = f"Enter an integer {low}-{high}"
    dv.errorTitle = "Invalid Index"
    ws.add_data_validation(dv)
    dv.add(cell)


def build_neuron_explorer_sheet(wb):
    """Sheet that visualizes any hidden neuron's 28x28 weight map."""
    ws = wb.create_sheet(title="Neuron Explorer")

    ws["A1"] = "Neuron Explorer — hidden unit weight map"
    ws["A2"] = "Neuron (1-32):"
    ws["B2"] = 1
    add_int_range_validation(ws, "B2", 1, 32)

    # Show current activation for selected neuron from Draw sheet
    ws["D2"] = "Activation (Draw!AI)"
    ws["E2"] = "=INDEX(Draw!$AI$1:$AI$32, $B$2)"

    # Weight grid A5:AB32 -> INDEX into Weights!A1:AF784 with pixel index k and selected neuron col
    top_row = 5
    left_col = 1  # A
    for r in range(28):
        for c in range(28):
            cell = ws.cell(row=top_row + r, column=left_col + c)
            # k = (r)*28 + (c+1) when r,c are 0-indexed; here: (ROW()-5)*28 + COLUMN()
            cell.value = "=INDEX(Weights!$A$1:$AF$784, (ROW()-5)*28 + COLUMN(), $B$2)"

    # Make cells square-ish
    for c in range(1, 28 + 1):
        ws.column_dimensions[excel_col_letter(c)].width = 2.5
    for r in range(top_row, top_row + 28):
        ws.row_dimensions[r].height = 14

    # Apply diverging color scale centered at 0
    add_diverging_scale(ws, "A5:AB32")

    ws.freeze_panes = "A5"
    return ws


def build_attention_sheet(wb, labels: List[str]):
    """Sheet that shows per-pixel contribution heat for the current Draw input."""
    ws = wb.create_sheet(title="Attention")

    ws["A1"] = "Per-pixel contribution heat (approx) for current prediction"
    ws["A2"] = "Predicted class index (1-5)"
    ws["B2"] = "=MATCH(MAX(Draw!$AF$2:$AF$6), Draw!$AF$2:$AF$6, 0)"
    ws["D2"] = "Predicted label"
    ws["E2"] = "=INDEX(Draw!$AE$2:$AE$6, $B$2)"

    # Attention formula: sum_j |W1[k,j]| * relu(h_j) * relu(W2[j, class])
    # Using SUMPRODUCT over vectors of length 32 with array tricks
    top_row = 5
    for r in range(28):
        for c in range(28):
            cell = ws.cell(row=top_row + r, column=1 + c)
            # k index formula uses sheet ROW/COLUMN
            # Use COLUMN(Weights!$A$1:$AF$1) to produce 1..32 explicitly and avoid implicit intersection (@) issues
            # Attention approx: sum_j |W1[k,j]| * relu(h_j) * relu(W2[j, class])
            formula = (
                "=SUMPRODUCT(" 
                "ABS(INDEX(Weights!$A$1:$AF$784, (ROW()-5)*28 + COLUMN(), COLUMN(Weights!$A$1:$AF$1)))," 
                "(Draw!$AI$1:$AI$32>0)*Draw!$AI$1:$AI$32," 
                "(INDEX(Weights!$A$790:$E$821, 0, $B$2)>0)*INDEX(Weights!$A$790:$E$821, 0, $B$2)" 
                ")"
            )
            cell.value = formula

    # Sizing and heatmap
    for c in range(1, 28 + 1):
        ws.column_dimensions[excel_col_letter(c)].width = 2.5
    for r in range(top_row, top_row + 28):
        ws.row_dimensions[r].height = 14

    add_red_heat(ws, "A5:AB32")
    ws.freeze_panes = "A5"
    return ws

def export_to_excel(W1, b1, W2, b2, labels: List[str], samples: List[Tuple[str, np.ndarray]], out_path: str):
    from openpyxl import Workbook

    print("[excel] Building workbook ...")
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)


    # Weights sheet first (hidden)
    write_weights_sheet(wb, W1, b1, W2, b2)

    # Main drawing sheet
    build_draw_sheet(wb, "Draw", labels=labels, sample_grid=None)

    # Exploratory/tech sheets
    build_neuron_explorer_sheet(wb)
    build_attention_sheet(wb, labels)


    print(f"[excel] Saving to {out_path} ...")
    wb.save(out_path)
    print("[excel] Saved.")


def save_model_json(W1, b1, W2, b2, labels: List[str], out_path: str):
    import json
    data = {
        "W1": np.asarray(W1, dtype=float).tolist(),
        "b1": np.asarray(b1, dtype=float).tolist(),
        "W2": np.asarray(W2, dtype=float).tolist(),
        "b2": np.asarray(b2, dtype=float).tolist(),
        "labels": list(labels),
    }
    with open(out_path, "w") as f:
        json.dump(data, f)
    print(f"[json] Saved model to {out_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Neural Network in Excel — QuickDraw Demo")
    parser.add_argument("--data-dir", default="data", help="Directory for .npy files")
    parser.add_argument("--out", default="NeuralNet_Excel_Demo2.xlsx", help="Output .xlsx path")
    parser.add_argument("--out-json", default=None, help="Optional model JSON output path (default: alongside --out)")
    parser.add_argument("--per-class", type=int, default=30000, help="Samples per class")
    parser.add_argument("--epochs", type=int, default=30, help="Max training epochs")
    parser.add_argument("--seed", type=int, default=0)
    args = parser.parse_args()

    # 1) Download data
    download_quickdraw(args.data_dir, max_per_class=args.per_class)

    # 2) Load data
    X, y, labels = load_dataset(args.data_dir, n_per_class=args.per_class)
    print(f"[data] Loaded X={X.shape}, y={y.shape}, classes={labels}")

    # 3) Train model
    model = train_model(X, y, max_epochs=args.epochs, target_acc=0.95, batch_size=256, lr=1e-3, val_split=0.15, seed=args.seed, min_epochs=10)

    # 4) Prepare sample grids (first from each class)
    samples = []
    # Load again non-shuffled to pick first sample per class
    for name, _ in CATEGORIES:
        arr = np.load(os.path.join(args.data_dir, f"{name}.npy"))
        grid = (arr[0].reshape(28, 28) > 0).astype(np.int64)  # 0/1
        samples.append((name, grid))

    # 5) Export to Excel + JSON (JSON path defaults next to Excel)
    export_to_excel(model.W1, model.b1, model.W2, model.b2, labels, samples, args.out)
    json_out = args.out_json or os.path.splitext(args.out)[0] + ".json"
    save_model_json(model.W1, model.b1, model.W2, model.b2, labels, json_out)

    # Final train/val report for console
    N = X.shape[0]
    n_val = int(N * 0.15)
    pred = model.predict(X[:n_val])
    acc = (pred == y[:n_val]).mean()
    print(f"[done] Validation accuracy: {acc*100:.2f}% | Output: {args.out}")


if __name__ == "__main__":
    # Lazy-import heavy deps where needed; openpyxl/requests imported in functions.
    main()
